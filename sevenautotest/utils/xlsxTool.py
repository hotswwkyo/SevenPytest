import re

from openpyxl import load_workbook, Workbook
import os


class xlsxTool(object):
    def __init__(self, filePath):
        self.filePath = filePath
        if os.path.exists(filePath):
            self.wb = load_workbook(filePath)
        else:
            self.wb = Workbook()

    @property
    def current_sheet(self):
        return self.wb.active

    @current_sheet.setter
    def current_sheet(self, value):
        self.wb.active = value

    def create_sheet(self, title):
        self.wb.create_sheet(title=title)
        self.wb.active = title

    def read_cell(self, row, column):
        if isinstance(row, int):
            row = str(row)
        if not (column, str):
            raise TypeError("Column must be string")
        target_cell = column + row
        cell_compile = re.compile(r"^[A-Z]+[0-9]+$")
        if re.match(cell_compile, target_cell):
            return self.current_sheet[target_cell]
        else:
            raise ValueError("Illegal Character {}".format(target_cell))

    @staticmethod
    def string_to_column_index(column):
        cell_compile = re.compile(r"^[A-Z]+$")
        if re.match(cell_compile, column):
            s_len = len(column)
            s_index = 0
            for i in range(s_len):
                index = s_len - i - 1
                c_num = ord(column[index]) - 64
                s_index += 26**i * c_num
            return s_index
        else:
            raise ValueError("Illegal Value {}".format(column))

    def write_cell(self, row, col, value):
        if isinstance(col, str):
            col = xlsxTool.string_to_column_index(col)
        if isinstance(row, str):
            row = int(row)
        self.current_sheet.cell(column=col, row=row, value=value)

    def save(self, filePath=None):
        if filePath:
            self.wb.save(filePath)
        else:
            self.wb.save(self.filePath)
